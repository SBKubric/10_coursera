import argparse
from bs4 import BeautifulSoup
import json
from lxml import etree
from concurrent import futures
import openpyxl
from random import sample
import requests

XML_FEED_URL = 'https://www.coursera.org/sitemap~www~courses.xml'
TITLES_LIST = ['title', 'language', 'startDate', 'weeks_counter', 'average_rating']
COURSES_AMOUNT = 20



def parse_args():
    parser = argparse.ArgumentParser(description='''This script fetches info about 20 random courses
                                                    from Coursera XML feed to XSL file.''')
    parser.add_argument('-o', '--output', default='courses.xlsx', help='''Name of the file with results.
                                                                          Default value is courses.xlsx''')
    return parser.parse_args()


def get_course_url_list():
    xml_page = requests.get(XML_FEED_URL)
    xml_data = etree.fromstring(xml_page.content)
    urls = [url.text for url in xml_data.iter('{*}loc')]
    courses_urls = sample(urls, COURSES_AMOUNT)
    return courses_urls


def get_course_data(course_url):
    request = requests.get(course_url)
    return request.content


def convert_soup_to_text(soup):
    return soup.text if soup else None


def get_course_date(soup):
    json_params_text = convert_soup_to_text(soup.find('script type="application/ld+json"'))
    if json_params_text:
        json_params = json.loads(json_params_text)
        return json_params['startDate']


def get_weeks_counter(soup):
    weeks_tag = soup.find_all('div', {'class': 'week'})
    return len(weeks_tag)


def get_average_rating(soup):
    soup_rating = soup.find('div', {'class': 'ratings-text bt3-visible-xs'})
    return soup_rating.text if soup_rating else None


def retrieve_courses_info(courses_urls):
    pool = futures.ThreadPoolExecutor(len(courses_urls))
    course_pages = list(pool.map(get_course_data, courses_urls))
    course_soups = [BeautifulSoup(course_page, 'html.parser') for course_page in course_pages]
    courses_info = list({
        'title': soup.find('div', {'class': 'title display-3-text'}).text,
        'language': soup.find('div', {'class': 'language-info'}).text,
        'startDate': get_course_date(soup),
        'weeks_counter': get_weeks_counter(soup),
        'average_rating': get_average_rating(soup),
    } for soup in course_soups)
    return courses_info


def output_courses_info_to_xlsx(course_slugs):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'courses_data'
    column_number = 0
    for title in TITLES_LIST:
        column_number += 1
        worksheet.cell(row=1, column=column_number).value = title
    for row_number, course_slug in enumerate(course_slugs, start=2):
        column_number = 0
        for title in TITLES_LIST:
            column_number += 1
            worksheet.cell(row=row_number, column=column_number).value = course_slug[title]
    return workbook


if __name__ == '__main__':
    args = parse_args()
    print('Getting courses urls...')
    courses_url_list = get_course_url_list()
    print('Collecting info about courses...')
    courses_info_list = retrieve_courses_info(courses_url_list)
    print('Creating xlsx file...')
    workbook = output_courses_info_to_xlsx(courses_info_list)
    workbook.save(args.output)
    print('Saved to {}'.format(args.output))
